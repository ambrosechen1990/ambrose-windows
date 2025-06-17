import tkinter as tk
from tkinter import ttk, filedialog, messagebox, Toplevel, simpledialog
import cv2
import os
from datetime import datetime
import numpy as np
import logging
import atexit
import zstandard as zstd
import shutil
import tempfile
from PIL import Image, ImageTk
import threading
import concurrent.futures
import sys
from tkinterdnd2 import DND_FILES, TkinterDnD
import multiprocessing
import zipfile
import tarfile
from concurrent.futures import ProcessPoolExecutor
import json
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
import openpyxl.styles
from openpyxl.utils import get_column_letter
import pytesseract
import re
import time
import subprocess
import webbrowser
import socket


def resource_path(relative_path):
    # 兼容pyinstaller打包和源码运行
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


def process_one_bin(args):
    src_file, target_dir, filename = args
    print(f"[DEBUG] 正在解析: {src_file}，输出到: {os.path.join(target_dir, filename[:-4] + '.log')}")
    log_file = os.path.join(target_dir, filename[:-4] + ".log")
    try:
        hex_list = []
        with open(src_file, 'r') as f_in:
            for line in f_in:
                line = line.strip()
                if len(line) > 14:
                    hex_list.append(line[14:])
        hex_str = ''.join(hex_list)
        hex_str = ''.join(filter(lambda c: c in '0123456789abcdefABCDEF', hex_str))
        data = bytes.fromhex(hex_str)
        zstd_magic = b'\x28\xb5\x2f\xfd'
        idx = 0
        with open(log_file, 'w', encoding='utf-8') as f_out:
            while True:
                idx = data.find(zstd_magic, idx)
                if idx == -1:
                    break
                next_idx = data.find(zstd_magic, idx + 4)
                chunk = data[idx:next_idx] if next_idx != -1 else data[idx:]
                try:
                    dctx = zstd.ZstdDecompressor()
                    decompressed = dctx.decompress(chunk)
                    f_out.write(decompressed.decode('utf-8', errors='replace'))
                except Exception as e:
                    print(f'解压第{idx}段失败: {e}')
                idx = next_idx if next_idx != -1 else len(data)
        print(f"已生成log文件: {log_file}")
        return 1
    except Exception as e:
        print(f"转换{src_file}失败: {e}")
        return 0


# 轨迹线绘制信息弹窗（含历史）
def get_history(path):
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_history(path, value):
    history = get_history(path)
    if value and value not in history:
        history.append(value)
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False)

def show_info_dialog():
    root = tk.Toplevel()
    root.title("填写轨迹线信息")
    sn_history = get_history('sn_history.json')
    pool_history = get_history('pool_history.json')
    fw_history = get_history('fw_history.json')
    tk.Label(root, text="机器序号:").grid(row=0, column=0)
    sn_var = tk.StringVar()
    sn_combo = ttk.Combobox(root, textvariable=sn_var, values=sn_history, width=30, font=("微软雅黑", 14))
    sn_combo.grid(row=0, column=1)
    tk.Label(root, text="泳池编号:").grid(row=1, column=0)
    pool_var = tk.StringVar()
    pool_combo = ttk.Combobox(root, textvariable=pool_var, values=pool_history, width=30, font=("微软雅黑", 14))
    pool_combo.grid(row=1, column=1)
    tk.Label(root, text="机器阶段:").grid(row=2, column=0)
    stage_var = tk.StringVar()
    stage_combo = ttk.Combobox(root, textvariable=stage_var, values=["手板","T0","EVT1","EVT2","DVT1","DVT2","MP"], width=30, font=("微软雅黑", 14))
    stage_combo.grid(row=2, column=1)
    tk.Label(root, text="固件版本号:").grid(row=3, column=0)
    fw_var = tk.StringVar()
    fw_combo = ttk.Combobox(root, textvariable=fw_var, values=fw_history, width=30, font=("微软雅黑", 14))
    fw_combo.grid(row=3, column=1)
    result = {}
    def on_ok():
        result['sn'] = sn_var.get()
        result['pool'] = pool_var.get()
        result['stage'] = stage_var.get()
        result['fw'] = fw_var.get()
        save_history('sn_history.json', result['sn'])
        save_history('pool_history.json', result['pool'])
        save_history('fw_history.json', result['fw'])
        root.destroy()
    tk.Button(root, text="确定", command=on_ok, font=("微软雅黑", 14)).grid(row=4, column=0, columnspan=2, pady=10)
    root.grab_set()
    root.wait_window()
    return result

def append_to_excel(info, img_path):
    dist_dir = r'D:/dist'
    os.makedirs(dist_dir, exist_ok=True)
    excel_path = os.path.join(dist_dir, '轨迹线绘制记录.xlsx')
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.append(['序号','视频开始时间','机器序号','泳池编号','机器阶段','固件版本号','绘制完成轨迹线地图','结束状态','覆盖率'])
        for cell in ws[ws.max_row]:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        row = [ws.max_row, info['start_time'], info['sn'], info['pool'], info['stage'], info['fw'], os.path.basename(img_path), info['end_status'], info['coverage']]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        img = XLImage(img_path)
        img.width = 200
        img.height = 150
        img.anchor = f'G{ws.max_row}'
        ws.add_image(img)
        col_letter = get_column_letter(7)
        ws.column_dimensions[col_letter].width = 35
        ws.row_dimensions[ws.max_row].height = 120
        while True:
            try:
                wb.save(excel_path)
                break
            except PermissionError:
                messagebox.showerror("保存失败", "Excel 文件已被打开，请关闭后点击确定重试。")
                time.sleep(1)
            except Exception as e:
                messagebox.showerror("保存失败", f"保存 Excel 时发生错误：{e}")
                break
    else:
        wb = load_workbook(excel_path)
        ws = wb.active
        row = [ws.max_row, info['start_time'], info['sn'], info['pool'], info['stage'], info['fw'], os.path.basename(img_path), info['end_status'], info['coverage']]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        img = XLImage(img_path)
        img.width = 200
        img.height = 150
        img.anchor = f'G{ws.max_row}'
        ws.add_image(img)
        col_letter = get_column_letter(7)
        ws.column_dimensions[col_letter].width = 35
        ws.row_dimensions[ws.max_row].height = 120
        while True:
            try:
                wb.save(excel_path)
                break
            except PermissionError:
                messagebox.showerror("保存失败", "Excel 文件已被打开，请关闭后点击确定重试。")
                time.sleep(1)
            except Exception as e:
                messagebox.showerror("保存失败", f"保存 Excel 时发生错误：{e}")
                break


class TrajectoryLine:
    def __init__(self):
        # 固定视频帧大小和默认轨迹线宽度
        self.FRAME_WIDTH = 640
        self.FRAME_HEIGHT = 480
        self.TRACK_WIDTH = 15  # 默认轨迹线宽度

        # 设置日志文件夹路径
        self.LOG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "log")
        os.makedirs(self.LOG_DIR, exist_ok=True)

        # 设置日志文件名称和路径
        log_file_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + ".log"
        log_file_path = os.path.join(self.LOG_DIR, log_file_name)

        # 配置日志记录
        logging.basicConfig(
            filename=log_file_path,
            level=logging.DEBUG,
            format="%(asctime)s - %(levelname)s - %(message)s"
        )
        atexit.register(logging.shutdown)

    def create_tracker(self):
        """创建跟踪器，兼容不同OpenCV版本"""
        try:
            if hasattr(cv2, 'legacy') and hasattr(cv2.legacy, 'TrackerCSRT_create'):
                return cv2.legacy.TrackerCSRT_create()
            elif hasattr(cv2, 'TrackerCSRT_create'):
                return cv2.TrackerCSRT_create()
            else:
                logging.error("未找到CSRT跟踪器")
                return None
        except Exception as e:
            logging.error(f"创建跟踪器失败: {str(e)}")
            return None

    def extract_time_from_frame(self, frame):
        h, w, _ = frame.shape
        roi = frame[h-60:h, w-250:w]  # 右下角区域，可根据实际调整
        pil_img = Image.fromarray(cv2.cvtColor(roi, cv2.COLOR_BGR2RGB))
        text = pytesseract.image_to_string(pil_img, config='--psm 7')
        match = re.search(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}', text)
        if match:
            return match.group(0)
        else:
            return ""

    def process_video(self, video_path, info):
        try:
            frame_count = 0
            coverage_rate = 0
            if not os.path.exists(video_path):
                logging.error(f"视频文件 {video_path} 不存在")
                return

            cap = cv2.VideoCapture(video_path)
            if not cap.isOpened():
                logging.error(f"无法打开视频文件 {video_path}")
                return

            ret, frame = cap.read()
            if not ret:
                logging.error("无法读取视频文件")
                return

            # 自动识别视频右下角时间
            video_time = self.extract_time_from_frame(frame)
            if video_time:
                info['start_time'] = video_time
            else:
                info['start_time'] = ""

            frame = cv2.resize(frame, (self.FRAME_WIDTH, self.FRAME_HEIGHT))

            tracker = None
            init_box = None
            all_track_points = []
            polygon_points = []  # 存储多边形的点
            drawing_polygon = True  # 标记是否在绘制多边形

            def on_mouse(event, x, y, flags, param):
                nonlocal drawing_polygon
                if drawing_polygon:
                    if event == cv2.EVENT_LBUTTONDOWN:
                        polygon_points.append((x, y))
                    elif event == cv2.EVENT_RBUTTONDOWN and len(polygon_points) > 2:
                        drawing_polygon = False

            cv2.namedWindow("Tracking")
            cv2.setMouseCallback("Tracking", on_mouse)
            print("请使用鼠标左键点击绘制多边形区域，右键完成绘制")

            # 绘制多边形区域
            while drawing_polygon:
                temp_frame = frame.copy()
                if len(polygon_points) > 1:
                    for i in range(1, len(polygon_points)):
                        cv2.line(temp_frame, polygon_points[i - 1], polygon_points[i], (0, 255, 255), 2)
                    if len(polygon_points) > 2:
                        cv2.line(temp_frame, polygon_points[-1], polygon_points[0], (0, 255, 255), 2)

                cv2.imshow("Tracking", temp_frame)
                key = cv2.waitKey(1) & 0xFF
                
                # 检查窗口是否被关闭
                if cv2.getWindowProperty("Tracking", cv2.WND_PROP_VISIBLE) < 1:
                    print("窗口被关闭，退出多边形绘制")
                    cv2.destroyAllWindows()
                    return
                    
                if key == ord('q') and len(polygon_points) > 2:
                    drawing_polygon = False

            if len(polygon_points) < 3:
                print("多边形区域无效，至少需要3个点")
                return

            # 创建多边形掩码
            mask = np.zeros((self.FRAME_HEIGHT, self.FRAME_WIDTH), dtype=np.uint8)
            cv2.fillPoly(mask, [np.array(polygon_points, np.int32)], 255)
            polygon_area = cv2.countNonZero(mask)

            # 找到多边形的轮廓
            contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            # 获取多边形内的所有点
            points_inside_polygon = []
            for y in range(self.FRAME_HEIGHT):
                for x in range(self.FRAME_WIDTH):
                    if cv2.pointPolygonTest(contours[0], (x, y), False) >= 0:
                        points_inside_polygon.append((x, y))

            white_trail = np.zeros((self.FRAME_HEIGHT, self.FRAME_WIDTH, 3), dtype=np.uint8)

            print("按空格键选择要跟踪的目标，按 q 键退出")
            end_status = 'Yes'
            
            while True:
                frame_count += 1
                if not ret:
                    print("视频播放完毕或读取失败")
                    break

                overlay = frame.copy()

                # 显示多边形区域
                cv2.polylines(overlay, [np.array(polygon_points, np.int32)], isClosed=True, color=(0, 255, 255), thickness=2)

                # 绘制轨迹线（透明绿色）
                for i in range(1, len(all_track_points)):
                    if all_track_points[i - 1] and all_track_points[i]:
                        cv2.line(overlay, all_track_points[i - 1], all_track_points[i], (0, 255, 0), self.TRACK_WIDTH)
                        cv2.line(white_trail, all_track_points[i - 1], all_track_points[i], (127, 127, 127), max(1, self.TRACK_WIDTH // 4))

                # 叠加白色轨迹层
                track_overlay = cv2.add(overlay, white_trail)

                if frame_count % 20 == 0:
                    covered_area = 0
                    for point in points_inside_polygon:
                        x, y = point
                        if overlay[y, x][1] == 255 and overlay[y, x][0] == 0 and overlay[y, x][2] == 0:
                            covered_area += 1
                    coverage_rate = (covered_area / polygon_area) * 100 if polygon_area > 0 else 0

                cv2.putText(overlay, f"Coverage: {coverage_rate:.2f}%", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 139, 255), 2)

                # 显示进度条
                total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
                current_frame = int(cap.get(cv2.CAP_PROP_POS_FRAMES))
                progress = current_frame / total_frames if total_frames > 0 else 0

                progress_bar_width = int(self.FRAME_WIDTH * progress)
                cv2.rectangle(overlay, (0, self.FRAME_HEIGHT - 10), (self.FRAME_WIDTH, self.FRAME_HEIGHT), (50, 50, 50), -1)
                cv2.rectangle(overlay, (0, self.FRAME_HEIGHT - 10), (progress_bar_width, self.FRAME_HEIGHT), (0, 255, 0), -1)

                # 显示结果帧
                alpha = 0.3
                result_frame = cv2.addWeighted(overlay, alpha, frame, 1 - alpha, 0)
                result_track_frame = cv2.addWeighted(track_overlay, alpha, frame, 1 - alpha, 0)
                cv2.imshow("Coverage", result_frame)
                cv2.imshow("Tracking", result_track_frame)

                key = cv2.waitKey(1) & 0xFF
                if key == ord('q') or key == ord('Q'):
                    end_status = 'No'
                    break
                elif key == ord(' '):
                    init_box = cv2.selectROI("Select object", frame, fromCenter=False)
                    if any(init_box):
                        tracker = self.create_tracker()
                        if tracker is not None:
                            tracker.init(frame, init_box)
                            current_track_points = []
                            all_track_points.extend(current_track_points)
                            print("目标选择完成，开始跟踪")
                        else:
                            print("无法初始化跟踪器，请确保已安装 OpenCV contrib 模块")
                    cv2.destroyWindow("Select object")

                if tracker:
                    success, bbox = tracker.update(frame)
                    if success:
                        x, y, w, h = [int(v) for v in bbox]
                        center_point = (int(x + w / 2), int(y + h / 2))
                        all_track_points.append(center_point)
                        # 在当前帧上显示跟踪框
                        cv2.rectangle(result_track_frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                    else:
                        print("目标跟踪失败，请重新选择目标")
                        tracker = None

                ret, frame = cap.read()
                if ret:
                    frame = cv2.resize(frame, (self.FRAME_WIDTH, self.FRAME_HEIGHT))

            # 保存最后一帧图片
            if len(all_track_points) > 0:
                output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "picture")
                os.makedirs(output_dir, exist_ok=True)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = os.path.join(output_dir, f"Coverage_rate_{timestamp}.png")
                # 使用PIL保存图片，设置dpi为300
                pil_img = Image.fromarray(cv2.cvtColor(result_frame, cv2.COLOR_BGR2RGB))
                pil_img.save(output_path, dpi=(300, 300))
                print(f"图像已保存至 {output_path}")
                # 写入Excel
                info['end_status'] = end_status
                info['coverage'] = f"{coverage_rate:.2f}%"
                append_to_excel(info, output_path)
            else:
                print("未检测到有效的轨迹线")

        except Exception as e:
            logging.error(f"处理视频时出现错误: {str(e)}")
            print(f"处理视频时出现错误: {str(e)}")
        finally:
            if 'cap' in locals():
                cap.release()
            cv2.destroyAllWindows()


class MainApplication:
    def __init__(self, root):
        print("进入MainApplication.__init__")
        self.root = root
        self.root.title("Beatbot软测工具")
        self.is_parsing = False  # 防抖标志

        # 设置窗口大小为720P并允许调整
        self.root.geometry("1280x720")
        self.root.minsize(1024, 576)

        # 配置根窗口的网格权重
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        # 设置样式
        self.setup_styles()

        # 创建主框架
        self.main_frame = ttk.Frame(root)
        self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=20, pady=20)

        # 配置主框架的网格权重
        for i in range(2):  # 2行
            self.main_frame.grid_rowconfigure(i, weight=1)
        for i in range(4):  # 4列
            self.main_frame.grid_columnconfigure(i, weight=1)

        # 创建轨迹线处理器实例
        self.trajectory = TrajectoryLine()

        # 初始化进度条相关变量
        self.progress_var = tk.DoubleVar()
        self.progress_bar = None
        self.progress_label = None
        self.progress_bottom = None

        # 创建功能区域
        self.create_function_areas()

    def setup_styles(self):
        style = ttk.Style()
        # 配置标签样式
        style.configure(
            'Icon.TLabel',
            font=('微软雅黑', 48),  # 大图标
            padding=10,
            anchor='center',  # 文本居中
            justify='center'  # 多行文本居中
        )
        style.configure(
            'Function.TLabel',
            font=('微软雅黑', 12, 'bold'),  # 功能名称字体
            padding=5,
            anchor='center',  # 文本居中
            justify='center'  # 多行文本居中
        )
        # 配置按钮样式
        style.configure(
            'Function.TButton',
            padding=10
        )

    def update_image_size(self, event, label, img_path):
        w, h = event.width, event.height
        max_img_w, max_img_h = int(w * 0.7), int(h * 0.5)
        try:
            img = Image.open(img_path)
            img = img.resize((max_img_w, max_img_h), Image.ANTIALIAS)
            photo = ImageTk.PhotoImage(img)
            label.config(image=photo)
            label.image = photo
        except Exception:
            pass

    def create_function_areas(self):
        """创建功能区域（卡片尺寸固定+进度条区始终占位）"""
        self.card_progress = {}
        function_cards = [
            {"name": "轨迹线绘制", "command": self.mcu_tools, "row": 0, "column": 0, "icon": resource_path("icons/轨迹线绘制.jpeg")},
            {"name": "日志解析", "command": self.unzip_and_parse_zip, "row": 0, "column": 1, "icon": resource_path("icons/日志解析.jpeg")},
            {"name": "日志打包下载", "command": self.pack_log, "row": 0, "column": 2, "icon": resource_path("icons/日志打包下载.jpeg")},
            {"name": "使用帮助", "command": self.show_help, "row": 0, "column": 3, "icon": resource_path("icons/使用帮助.jpeg")},
        ]
        for i in range(2):
            self.main_frame.grid_rowconfigure(i, weight=1)
        for i in range(4):
            self.main_frame.grid_columnconfigure(i, weight=1)
        for row in range(2):
            for col in range(4):
                func = next((f for f in function_cards if f["row"] == row and f["column"] == col), None)
                frame = ttk.Frame(
                    self.main_frame,
                    relief='solid',
                    borderwidth=1,
                    width=260,
                    height=220
                )
                frame.grid(
                    row=row,
                    column=col,
                    sticky="nsew",
                    padx=16,
                    pady=16
                )
                frame.grid_propagate(False)
                frame.pack_propagate(False)
                # 内容区
                content = ttk.Frame(frame)
                content.pack(expand=True, fill='both')
                # 进度条区（始终占位）
                progress_area = ttk.Frame(frame, height=40)
                progress_area.pack(fill='x', side='bottom')
                progress_area.pack_propagate(False)
                if func:
                    icon_img = None
                    if func["icon"] and os.path.exists(func["icon"]):
                        try:
                            img = Image.open(func["icon"])
                            img = img.resize((64, 64), Image.ANTIALIAS)
                            icon_img = ImageTk.PhotoImage(img)
                        except Exception:
                            pass
                    label = ttk.Label(content, text=func["name"], style='Function.TLabel', image=icon_img, compound='top')
                    label.image = icon_img
                    label.pack(expand=True, fill='both', pady=(20, 0))
                    label.bind("<Button-1>", lambda e, f=func["command"]: f())
                    # 独立进度条和标签
                    progress_var = tk.DoubleVar()
                    progress_bar = ttk.Progressbar(progress_area, variable=progress_var, length=180, mode='determinate')
                    progress_label = ttk.Label(progress_area, text="", font=("微软雅黑", 10))
                    progress_bar.pack(side='top', fill='x', padx=20, pady=(5, 0))
                    progress_label.pack(side='top', fill='x', padx=20, pady=(0, 10))
                    progress_bar.pack_forget()
                    progress_label.pack_forget()
                    self.card_progress[(row, col)] = {
                        'bar': progress_bar,
                        'label': progress_label,
                        'var': progress_var
                    }

    def show_card_progress(self, row, col, total):
        p = self.card_progress.get((row, col))
        if p:
            p['bar'].config(maximum=total)
            p['var'].set(0)
            p['bar'].pack(side='top', fill='x', padx=20, pady=(5, 0))
            p['label'].pack(side='top', fill='x', padx=20, pady=(0, 10))

    def update_card_progress(self, row, col, value, total, text=None):
        p = self.card_progress.get((row, col))
        if p:
            p['bar'].config(maximum=total)
            p['var'].set(value)
            if text:
                p['label'].config(text=text)
            else:
                percent = int((value / total) * 100)
                p['label'].config(text=f"进度：{percent}%")

    def close_card_progress(self, row, col):
        p = self.card_progress.get((row, col))
        if p:
            p['bar'].pack_forget()
            p['label'].pack_forget()

    def _make_card_command(self, cmd):
        return lambda e: cmd()

    def mcu_tools(self):
        threading.Thread(target=self._mcu_tools_impl, daemon=True).start()

    def _mcu_tools_impl(self):
        info = show_info_dialog()  # Show info dialog first
        if info:
            video_path = filedialog.askopenfilename(
                title="选择视频文件",
                filetypes=[
                    ("MP4 文件", "*.mp4"),
                    ("AVI 文件", "*.avi"),
                    ("MOV 文件", "*.mov"),
                    ("MKV 文件", "*.mkv"),
                    ("所有文件", "*.*")
                ]
            )
            if video_path:
                self.trajectory.process_video(video_path, info)

    def show_progress(self, total):
        if not hasattr(self, 'progress_var'):
            self.progress_var = tk.DoubleVar()
        if not hasattr(self, 'progress_bar') or self.progress_bar is None:
            self.progress_bar = ttk.Progressbar(self.main_frame, maximum=total, variable=self.progress_var, length=400)
            self.progress_bar.grid(row=2, column=0, columnspan=4, sticky='ew', padx=20, pady=(10, 0))
        if not hasattr(self, 'progress_label') or self.progress_label is None:
            self.progress_label = ttk.Label(self.main_frame, text="", font=("微软雅黑", 12))
            self.progress_label.grid(row=3, column=0, columnspan=4, sticky='ew', padx=20)
        self.progress_bar.config(maximum=total)
        self.progress_var.set(0)
        self.progress_bar.grid()
        self.progress_label.grid()

    def update_progress(self, value, total):
        if hasattr(self, 'progress_bar') and self.progress_bar:
            self.progress_bar.config(maximum=total)
            self.progress_var.set(value)
            self.progress_bar.update()
        if hasattr(self, 'progress_label') and self.progress_label:
            self.progress_label.update()

    def close_progress(self):
        if hasattr(self, 'progress_bar') and self.progress_bar:
            self.progress_bar.grid_remove()
        if hasattr(self, 'progress_label') and self.progress_label:
            self.progress_label.grid_remove()

    def show_help(self):
        if hasattr(self, 'help_win') and self.help_win and self.help_win.winfo_exists():
            self.help_win.lift()
            return
        self.help_win = tk.Toplevel(self.root)
        self.help_win.title("使用帮助")
        self.help_win.geometry("520x400")
        self.help_win.resizable(False, False)
        help_text = (
            "【环境依赖项】\n"
            "1. 安装 Tesseract-OCR，安装后需要在path中配置环境。\n"
            "2. 将C:\\Program Files\\Tesseract-OCR复制粘贴至path中后确定。\n"
            "3. 安装VC_redist.x64.exe，一直下一步。\n\n"
            "【轨迹线绘制】\n"
            "1. 点击'轨迹线绘制'卡片，填写机器序号、泳池编号、阶段、固件版本号等信息。\n"
            "2. 选择视频文件。\n"
            "3. 用鼠标左键依次点击视频画面，绘制多边形区域，右键闭合。\n"
            "4. 按空格选择跟踪目标，目标跟踪后会显示轨迹线和覆盖率。\n"
            "5. 按q或Shift+Q手动结束，结束状态为No，视频播放完毕自动结束为Yes。\n"
            "6. 轨迹线绘制信息（含图片、覆盖率、结束状态等）会自动写入Excel表格，图片自动缩放嵌入单元格。\n"
            "7. 需要本机已安装Tesseract-OCR（并配置到PATH），否则无法识别视频时间。\n"
            "8. openpyxl依赖已集成打包，无需单独安装。源码运行需pip install openpyxl。\n\n"
            "【文件解析】\n"
            "1. 点击'文件解析'卡片，可选择zip、tar.gz或tar格式的压缩包，自动解压并解析所有bin文件。\n"
            "2. 解析进度通过进度条显示，全部完成后弹窗提示解析数量。\n"
            "3. 解析生成的log文件与bin文件在同一目录，支持多层文件夹结构。"
        )
        text = tk.Text(self.help_win, wrap="word", font=("微软雅黑", 12), padx=10, pady=10)
        text.insert("1.0", help_text)
        text.config(state="disabled")
        text.pack(expand=True, fill="both", padx=10, pady=10)

    def unzip_and_parse_zip(self):
        archive_path = filedialog.askopenfilename(
            title="选择压缩包",
            filetypes=[
                ("压缩包", "*.zip *.tar.gz *.tar"),
                ("Zip files", "*.zip"),
                ("Tar GZ files", "*.tar.gz"),
                ("Tar files", "*.tar"),
                ("所有文件", "*.*")
            ]
        )
        if archive_path:
            threading.Thread(target=lambda: self.extract_zip_and_parse_with_progress(archive_path), daemon=True).start()

    def extract_zip_and_parse_with_progress(self, archive_path):
        import tarfile, zipfile, os
        extract_dir = os.path.splitext(os.path.splitext(archive_path)[0])[0] if archive_path.endswith('.tar.gz') else \
            os.path.splitext(archive_path)[0]
        print(f"[DEBUG] 解压目录: {extract_dir}")
        if archive_path.endswith('.zip'):
            with zipfile.ZipFile(archive_path, 'r') as zip_ref:
                zip_ref.extractall(extract_dir)
        elif archive_path.endswith('.tar.gz') or archive_path.endswith('.tar'):
            with tarfile.open(archive_path, 'r:*') as tar_ref:
                tar_ref.extractall(extract_dir)
        else:
            self.root.after(0, lambda: messagebox.showerror("错误", "不支持的压缩包格式"))
            return
        # 收集所有bin文件
        bin_files = []
        for root, dirs, files in os.walk(extract_dir):
            for file in files:
                if file.lower().endswith('.bin'):
                    bin_files.append((os.path.join(root, file), root, file))
        print(f"[DEBUG] 查找到bin文件: {bin_files}")
        total = len(bin_files)
        self.root.after(0, lambda: self.show_card_progress(0, 1, total))
        count = 0
        with ProcessPoolExecutor(max_workers=4) as executor:
            for result in executor.map(process_one_bin, bin_files):
                count += result
                self.root.after(0, lambda c=count: self.update_card_progress(0, 1, c, total, f"解析中... ({c}/{total})"))
        self.root.after(0, self.close_card_progress(0, 1))
        self.root.after(0, lambda: messagebox.showinfo("完成", f"共解析了 {count} 个 bin 文件"))

    def batch_convert_multi_folders(self, folders):
        import os
        import shutil
        import concurrent.futures

        if not folders:
            print("[DEBUG] 没有拖拽到任何文件夹")
            return

        # 防止重复运行
        if getattr(self, "_is_running_multi_folder", False):
            print("[DEBUG] 多文件夹转换任务已在运行中，忽略本次请求。")
            return
        self._is_running_multi_folder = True
        self._has_shown_multi_folder_msg = False

        # 1. 收集所有 bin 文件及其目标路径
        all_bin_files = []
        for folder in folders:
            new_folder = folder + "_log"
            for root, dirs, files in os.walk(folder):
                rel_path = os.path.relpath(root, folder)
                target_dir = os.path.join(new_folder, rel_path) if rel_path != '.' else new_folder
                os.makedirs(target_dir, exist_ok=True)
                for filename in files:
                    src_file = os.path.join(root, filename)
                    if filename.lower().endswith('.bin'):
                        all_bin_files.append((src_file, target_dir, filename))
                    else:
                        dst_file = os.path.join(target_dir, filename)
                        shutil.copy2(src_file, dst_file)

        total = len(all_bin_files)
        print(f"[DEBUG] 拖拽解析，总共 {total} 个 bin 文件")
        self.root.after(0, lambda: self.show_card_progress(0, 1, total))

        def run_and_update():
            count = 0
            max_workers = min(32, os.cpu_count() * 3)
            with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
                futures = [executor.submit(process_one_bin, args) for args in all_bin_files]
                for i, fut in enumerate(concurrent.futures.as_completed(futures), 1):
                    try:
                        result = fut.result()
                        count += result
                    except Exception as e:
                        print(f"[ERROR] 子任务失败: {e}")
                    self.root.after(0, lambda i=i: self.update_card_progress(0, 1, i, total, f"解析中... ({i}/{total})"))

            def show_msg():
                if self._has_shown_multi_folder_msg:
                    return
                self._has_shown_multi_folder_msg = True
                self.close_card_progress(0, 1)
                self.progress_label.config(
                    text=f"已将 {count} 个 bin 文件转为明文 log，其他文件已原样保留到各自 _log 文件夹"
                )
                self.is_parsing = False
                self._is_running_multi_folder = False
                print("[DEBUG] 多文件夹转换完成，弹窗提示")
                messagebox.showinfo("完成",
                                    f"已将 {count} 个 bin 文件转为明文 log，其他文件已原样保留到各自 _log 文件夹")

            self.root.after(0, show_msg)

        threading.Thread(target=run_and_update, daemon=True).start()

    def is_same_lan(self, ip):
        try:
            local_ip = socket.gethostbyname(socket.gethostname())
            return '.'.join(local_ip.split('.')[:3]) == '.'.join(ip.split('.')[:3])
        except:
            return False

    def pack_log(self):
        """日志打包下载功能"""
        ip = simpledialog.askstring("输入", "请输入设备IP：")
        if not ip:
            return
        if not self.is_same_lan(ip):
            messagebox.showerror("网络错误", "不在同一局域网内，请检查IP！")
            return
        # 显示进度条（使用日志打包下载卡片的进度条）
        self.show_card_progress(0, 2, 100)
        self.update_card_progress(0, 2, 0, 100, "打包中...")
        def do_pack():
            try:
                # 1. adb connect
                self.update_card_progress(0, 2, 0, 100, "正在连接设备...")
                connect_proc = subprocess.run(f'adb connect {ip}:5555', shell=True, capture_output=True, text=True, timeout=10)
                if ("connected" not in connect_proc.stdout) and ("already connected" not in connect_proc.stdout):
                    self.root.after(0, lambda: [self.close_card_progress(0, 2), messagebox.showerror("连接失败", connect_proc.stdout+connect_proc.stderr)])
                    return
                # 2. 执行pack命令
                self.update_card_progress(0, 2, 10, 100, "执行pack命令...")
                pack_proc = subprocess.run(f'adb shell pack', shell=True, capture_output=True, text=True, timeout=300)
                print("[PACK_CMD_OUT]", pack_proc.stdout)
                print("[PACK_CMD_ERR]", pack_proc.stderr)
                # 3. 查找日志包
                self.update_card_progress(0, 2, 60, 100, "打包完成，准备下载日志包...")
                ls_cmd = "adb shell ls /data/"
                ls_proc = subprocess.run(ls_cmd, shell=True, capture_output=True, text=True)
                print("[DEBUG] adb shell ls /data/ 输出：")
                print(ls_proc.stdout)
                ansi_escape = re.compile(r'\x1b\[[0-9;]*m')
                log_files = []
                pattern = re.compile(r"^manual_pack-.*\.tar\.gz$")
                for f in ls_proc.stdout.splitlines():
                    fname = ansi_escape.sub('', f.strip())
                    if not fname:
                        continue
                    print(f"[DEBUG] 文件名: '{fname}' repr: {repr(fname)} len: {len(fname)}")
                    if pattern.match(fname):
                        log_files.append(fname)
                print(f"[DEBUG] 匹配到的日志包文件: {log_files}")
                if not log_files:
                    print("[ERROR] 未找到日志包，全部文件名如下：")
                    for f in ls_proc.stdout.splitlines():
                        print(f"  [ALL FILE] '{f}'")
                    self.root.after(0, lambda: [self.close_card_progress(0, 2), messagebox.showerror("未找到日志包", "设备未生成日志包，请检查pack脚本")])
                    return
                # 4. 下载日志包
                remote_log_file = f"/data/{log_files[0]}"
                local_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')
                os.makedirs(local_path, exist_ok=True)
                self.update_card_progress(0, 2, 80, 100, "正在下载日志包...")
                pull_cmd = f'adb pull {remote_log_file} {local_path}'
                proc = subprocess.Popen(pull_cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
                out, _ = proc.communicate()
                if proc.returncode != 0:
                    self.root.after(0, lambda: [self.close_card_progress(0, 2), messagebox.showerror("下载失败", out.decode(errors='ignore'))])
                    return
                self.update_card_progress(0, 2, 100, 100, "日志打包并下载完成")
                self.root.after(0, lambda: [self.close_card_progress(0, 2), messagebox.showinfo("完成", f"日志已下载至: {local_path}")])
            except Exception as e:
                self.root.after(0, lambda: [self.close_card_progress(0, 2), messagebox.showerror("异常", f"日志打包流程异常：{e}")])
        threading.Thread(target=do_pack, daemon=True).start()


# 保证主入口只在主进程执行，防止多进程时重复启动GUI
if __name__ == "__main__":
    import traceback
    print("程序已启动")
    try:
        multiprocessing.freeze_support()  # 兼容 pyinstaller 多进程打包
        print("准备初始化TkinterDnD")
        root = TkinterDnD.Tk()
        print("TkinterDnD初始化完成")
        app = MainApplication(root)
        print("MainApplication初始化完成")
        root.mainloop()
    except Exception as e:
        print("程序启动异常：", e)
        traceback.print_exc()
        input("按回车键退出...")
